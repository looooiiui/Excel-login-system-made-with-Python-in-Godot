#====引入的 Openpyxl 库中的工作簿，工作表，以及加载函数====#
from openpyxl.worksheet.worksheet import Worksheet        # 工作表类
from openpyxl.workbook.workbook import Workbook           # 工作簿类
from openpyxl import load_workbook                        # 加载Excel函数
from typing import Optional                               # 多类型注释，用于函数返回多类型

#====允许调试信息(防止 print() 输出到Godot捕获参数列表)====#
DEBUG = False
# 调试信息输出
def log(msg):
    """只有 DEBUG=True 时才打印，否则不输出任何东西"""
    if DEBUG:
        print(msg)
#====允许调试信息(防止 print() 输出到Godot捕获参数列表)====#

"""
Excel全局Excel文件管理器
可以替换所有函数内打开Excel的操作
减少频繁打开关闭Excel的性能消耗
这里可以忽略，本文并未使用
"""
#===============全局文件管理器Excel===============#
"""
统一管理全局被打开的Excel
防止手动释放出现问题
"""
class Excel_Manager:

    #==默认空表未打开Excel==#
    def __init__(self):
        self.inner_excel: Workbook      = None

    #==尝试打开Excel==#
    def open_excel(self, file_path: str) -> bool:     
        #==检查本对象是否已经有Excel指向==#
        if self.inner_excel:
            log(f"打开对象已存在: {self.inner_excel}: 请手动释放")
            return False

        #==尝试读取Excel==#
        try:
            #==获取Excel==#
            read_excel: Workbook = load_workbook(file_path, read_only=True)
            self.inner_excel = read_excel
            return True
    
        #==异常处理==#
        except Exception as e:
            log(f"打开Excel发生异常: {e}")
            return False

    #==获得类中Excel==#
    def get_excel(self) -> Optional[Workbook]:
        #==检查本对象Excel是否已经指向==#
        if self.inner_excel:
            log(f"成功返回Excel: {self.inner_excel}")
            return self.inner_excel
        
        #=不存在=#
        log("返回excel失败,excel并未读取")
        return None

    #==获得Excel中对应表单==#   
    def get_sheet(self, sheet_name: str) -> Optional[Worksheet]:
        #==检查Excel指向==#
        if not self.inner_excel:
            log("不存在原表")
            return None

        #==返回特定表==#     
        try:
            if sheet_name not in self.inner_excel.sheetnames:
                log(f"工作表'{sheet_name}' 不存在")
                return None
            #==返回表==#
            return self.inner_excel[sheet_name]
        #==返回异常==#
        except Exception as e:
            log(f"表单返回发生异常: {e}")
            return None

    #==关闭Excel表格以及停止指向==#
    def close_workbook(self) -> bool:
        #==检查Excel指向==#
        if not self.inner_excel:
            log("不存在原表,已经关闭")
            return True
        try:
            self.inner_excel.close()
            return True
        
        #==表单关闭失败==#
        except Exception as e:
            log(f"关闭excel出现异常: {e}")
            return False   
        
"""
以下为当前文章使用的原版表格函数
因为没有Excel统一管理类
所以需要注意表格的状态
开启或者关闭
"""
#============从excel中以固定格式读取工作表============#
#===========返回值为 Worksheet 或 None ===============#
#==参数(Excel路径，工作表名字，是否只读(默认不只读))===#
def read_from_excel(file_read_path: str,
                    read_data_name: str,
                    read_only_out: bool = False) -> Optional[Worksheet]:
    
    read_excel: Workbook
    #==返回指定工作表==#
    try:
        #==获取Excel==#
        read_excel = load_workbook(file_read_path, read_only=read_only_out)
        #==Excel中不存在指定工作表==#
        if read_data_name not in read_excel.sheetnames:
            log(f"工作表'{read_data_name}' 不存在")
            read_excel.close()
            return None
        
        #==读取返回指定工作表==#
        read_excel_sheet = read_excel[read_data_name]
        read_excel.close()
        return read_excel_sheet
    
    # 读取失败
    except Exception as e:
        log(f"read failed{e}")
        # 打开了就关闭Excel
        if read_excel:
            read_excel.close()
        return None
    
"""
账户注册处理
返回-1: 程序运行出错
返回0 : 注册成功
返回1 : 账户注册出现问题(非法字符/账户密码问题)
返回2 : 账户已被注册
"""    
#=============写入新玩家数据==============#
#===返回值为元组 tuple[bool, int] ========#
#====第一个为是否成功，第二个为结果参数===#

#参数(Excel工作表名，Excel路径，新的账户信息)#
def write_new_account(read_data_name: str,
                      file_read_path: str,
                      new_account_information: Optional[list[str]]) -> tuple[bool, int]:

    #===传入空列表处理===#
    if new_account_information is None:
        return False, -1

    #=======这里调用了下面的验证账户输入合法性质(检验非法字符)检查函数=========#
    #==================返回同样为元组 tuple[bool, str] =======================#
    #==================== 第一个为是否合法，第二个为调试信息===================#
    get_account_verify: tuple[bool, str] = verify_account_data(new_account_information)

    #===非法返回 1===#
    if not get_account_verify[0]:
        log("账号登录或注册出现问题: " + get_account_verify[1])
        return False, 1

    read_excel: Workbook
    #=====合法尝试修改表格=====#
    try:
        #=====这里调用了下面验证Excel中玩家是否存在的函数=====#
        #==================返回值为bool类型==================#
        #=== 参数 (Excel工作表名字，Excel路径，查找账户ID) ===#
        if detect_id_excel_exist(read_data_name, file_read_path, new_account_information[0]):
            log(f"玩家ID已经存在: {new_account_information[0]}")
            return False, 2
    
        #===修改Excel表格，则Excel读取不能启用 read_only=True ===#
        read_excel = load_workbook(file_read_path, read_only=False)

        # 检查工作表存在
        if read_data_name not in read_excel.sheetnames:
            log(f"工作表'{read_data_name}' 不存在")
            read_excel.close()
            return False, -1
        
        # 选中工作表，获得新数据行号
        read_excel_sheet = read_excel[read_data_name]
        new_row_index = read_excel_sheet.max_row + 1        # 获得最大行的下一行，就是插入行
        col_num = read_excel_sheet.max_column               # 获得Excel当前工作表最大列数

        #====输入数据不全全部补"None"=====#
        need_add_num = col_num - len(new_account_information)       # 要补充的 "None" 的数量
        if need_add_num > 0:
            new_account_information += [None] * need_add_num

        #=======遍历新行所有单元格写入=======#
        # col_index 从 start 开始，每增加一列增加 1，用来计数
        # .cell()的第三个参数用来设置Excel单个格子的值
        for col_index, input_value in enumerate(new_account_information, start=1):
            read_excel_sheet.cell(row=new_row_index, 
                                  column=col_index, 
                                  value=input_value)

        #====保存修改数据====#
        read_excel.save(file_read_path)
        read_excel.close()
        log(f"数据写入成功: {new_account_information} -> {file_read_path}")
        return True, 0


    # 写入失败处理
    except Exception as e:
        log(f"write failed: {e}")
        if read_excel:
            read_excel.close()
        return False, -1

#=====验证Excel中玩家是否存在=====#
#=========返回值为bool类型========#
#=== 参数 (Excel工作表名字，Excel路径，查找账户ID) ===#
def detect_id_excel_exist(read_data_name: str, file_read_path: str, detected_id: str) -> bool:

    read_excel: Workbook
    #====尝试寻找ID====#
    try:
        # read_only=False，这样才可以使用列迭代器，否则报错
        read_excel = load_workbook(file_read_path, read_only=False)
        # 检查工作表是否存在于当前 Excel 表格
        if read_data_name not in read_excel.sheetnames:
            log(f"工作表'{read_data_name}' 不存在")
            read_excel.close()
            return False
        
        #====选中工作表，获取最大行数====#
        read_excel_sheet = read_excel[read_data_name]
        read_sheet_first = read_excel_sheet["A"]
        row_max_len = read_excel_sheet.max_row

        #====遍历确认玩家ID不存在====#
        exist_dir: dict[str, int] = {}
        for current_row, cell in enumerate(read_sheet_first, start=1):
            # 略过第一个属性名
            if current_row == 1:
                continue
            # 将所有玩家ID放入字典，用于后续检测
            if cell.value:
                exist_dir[str(cell.value)] = 1

        # 对比检测是否存在
        if detected_id in exist_dir:
            read_excel.close()
            return True
        read_excel.close()
        return False

    #=====检测出现异常，用于调试======#
    except Exception as e:
        log(f"查找Excel出现问题: {e}")
        if read_excel:
            read_excel.close()
        return False
    
#=======验证账户输入合法性质(检验非法字符)检查函数=========#
#===========返回为元组 tuple[bool, str] ==================#
#========== 第一个为是否合法，第二个为调试信息=============#
def verify_account_data(acc_info: list[str]) -> tuple[bool, str]:
    """
    账号合法性质校验、
    校验至少[账号ID, 密码, ......]
    """
    if not acc_info or len(acc_info) < 2:
        return False, "字段不足: 至少需要[账号ID, 密码]"
    
    acc_id          = acc_info[0]
    acc_password    = acc_info[1]

    #=====账号基本信息校验=====#
    # 返回 bool 类型
    if not acc_id:
        return False, "账号ID不能为空或全空格"
    if len(acc_id) < 3:
        return False, "账号长度不能小于3"
    # 这里右边 per_char 会遍历组成一个表格，只有非法字符会成为元素
    # any() 只要其中检测的元素不为空则返回true，也就是有非法字符返回true
    if any(per_char in r'\/:*?"<>| ' for per_char in acc_id):
        return False, "账号ID包含非法特殊字符"
    
    #====密码校验====#
    if not acc_password:
        return False, "密码不能为空"
    if len(acc_password) < 4:
        return False, "密码长度不能小于4"
    
    return True, "校验通过"